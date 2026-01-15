from fastapi import APIRouter, Request, Form, Depends, UploadFile, File, Query, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_
import os
import shutil
import database
import models
import auth
import excel_import
import pandas as pd
import io
import calendar
from datetime import datetime, date, timedelta
from typing import Optional
from models import TipoUsuario, EstadoProspecto
from dependencies import require_admin, get_current_user
from utils import (
    parsear_fecha,
    calcular_rango_fechas,
    normalizar_fecha_input,
    normalizar_texto_mayusculas,
    normalizar_numero,
    normalizar_email,
    enviar_notificacion_email
)

router = APIRouter()
templates = Jinja2Templates(directory="templates")

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ========== HELPER FUNCTIONS ==========

def crear_notificaciones_viaje(prospecto: models.Prospecto, db: Session):
    """
    Crea notificaciones automáticas para seguimiento de viaje.
    """
    # Validar que tenga fecha_ida
    if not prospecto.fecha_ida:
        return

    # Validar que sea estado ganado
    if prospecto.estado != EstadoProspecto.GANADO.value:
        return

    hoy = datetime.now().date()

    # Eliminar notificaciones automáticas anteriores de este prospecto
    db.query(models.Notificacion).filter(
        models.Notificacion.prospecto_id == prospecto.id,
        models.Notificacion.tipo == 'seguimiento_viaje'
    ).delete()

    # Definir notificaciones
    notificaciones_config = [
        {
            'dias_antes': 45,
            'mensaje': f'Confirmar pagos y estado de reserva - Viaje a {prospecto.destino or "destino"}'
        },
        {
            'dias_antes': 10,
            'mensaje': f'Validar con cliente gestiones pre-viaje - Viaje a {prospecto.destino or "destino"}'
        },
        {
            'dias_antes': 2,
            'mensaje': f'Validar pre-viaje, formularios y gestiones finales - Viaje a {prospecto.destino or "destino"}'
        }
    ]

    # Crear notificaciones
    for config in notificaciones_config:
        fecha_notificacion = prospecto.fecha_ida - timedelta(days=config['dias_antes'])

        # Solo crear si la fecha es futura
        if fecha_notificacion >= hoy:
            nueva_notificacion = models.Notificacion(
                prospecto_id=prospecto.id,
                usuario_id=prospecto.agente_asignado_id,
                tipo='seguimiento_viaje',
                mensaje=config['mensaje'],
                fecha_programada=datetime.combine(fecha_notificacion, datetime.min.time()),
                leida=False,
                email_enviado=False
            )
            db.add(nueva_notificacion)

    db.commit()

def obtener_ruta_upload_por_fecha(fecha: datetime = None) -> str:
    """
    Genera la ruta de directorio para uploads basada en la fecha.
    """
    if fecha is None:
        fecha = datetime.now()

    # Crear estructura: uploads/YYYY/MM/DD/
    ruta = os.path.join(
        UPLOAD_DIR,
        str(fecha.year),
        f"{fecha.month:02d}",
        f"{fecha.day:02d}"
    )

    # Crear directorios si no existen
    os.makedirs(ruta, exist_ok=True)

    return ruta

def generar_excel_prospectos(prospectos, filename="prospectos_export.xlsx"):
    """
    Genera archivo Excel con lista de prospectos
    """
    try:
        # Preparar datos para DataFrame
        data = []
        for p in prospectos:
            # Obtener última interacción
            ultima_interaccion = ""
            if p.interacciones:
                ultima_int = p.interacciones[0]  # Ya está ordenado por fecha desc
                ultima_interaccion = ultima_int.fecha_creacion.strftime("%d/%m/%Y %H:%M")

            data.append({
                'ID Cliente': p.id_cliente or f"CL-{p.id:04d}",
                'Nombre': p.nombre,
                'Apellido': p.apellido,
                'Teléfono': p.telefono or "",
                'Teléfono Secundario': p.telefono_secundario or "",
                'Email': p.correo_electronico or "",
                'Ciudad Origen': p.ciudad_origen or "",
                'Destino': p.destino or "",
                'Fecha Ida': p.fecha_ida.strftime("%d/%m/%Y") if p.fecha_ida else "",
                'Fecha Vuelta': p.fecha_vuelta.strftime("%d/%m/%Y") if p.fecha_vuelta else "",
                'Adultos': p.pasajeros_adultos or 0,
                'Niños': p.pasajeros_ninos or 0,
                'Infantes': p.pasajeros_infantes or 0,
                'Medio Ingreso': p.medio_ingreso.nombre if p.medio_ingreso else "",
                'Agente Asignado': p.agente_asignado.username if p.agente_asignado else "Sin asignar",
                'Estado': p.estado.replace("_", " ").title(),
                'Fecha Registro': p.fecha_registro.strftime("%d/%m/%Y %H:%M"),
                'Última Interacción': ultima_interaccion,
                'Cliente Recurrente': "Sí" if p.cliente_recurrente else "No",
                'Datos Completos': "Sí" if p.tiene_datos_completos else "No",
                'Observaciones': p.observaciones or ""
            })

        # Crear DataFrame
        df = pd.DataFrame(data)

        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Prospectos', index=False)

            # Obtener workbook y worksheet para formato
            workbook = writer.book
            worksheet = writer.sheets['Prospectos']

            # Formato de encabezados
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-ajustar columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Agregar filtros automáticos
            worksheet.auto_filter.ref = worksheet.dimensions

        output.seek(0)
        return output

    except Exception as e:
        print(f"❌ Error generando Excel de prospectos: {e}")
        import traceback
        traceback.print_exc()
        return None

def generar_excel_interacciones(interacciones, prospecto, filename="interacciones_export.xlsx"):
    """
    Genera archivo Excel con historial de interacciones
    """
    try:
        data = []
        for interaccion in interacciones:
            data.append({
                'Fecha': interaccion.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
                'Usuario': interaccion.usuario.username if interaccion.usuario else "Sistema",
                'Tipo': interaccion.tipo_interaccion.replace("_", " ").title() if interaccion.tipo_interaccion else "General",
                'Descripción': interaccion.descripcion,
                'Estado Anterior': interaccion.estado_anterior.replace("_", " ").title() if interaccion.estado_anterior else "",
                'Estado Nuevo': interaccion.estado_nuevo.replace("_", " ").title() if interaccion.estado_nuevo else ""
            })

        df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Información del prospecto
            info_data = {
                'Campo': ['ID Cliente', 'Nombre', 'Teléfono', 'Email', 'Estado Actual'],
                'Valor': [
                    prospecto.id_cliente or f"CL-{prospecto.id:04d}",
                    f"{prospecto.nombre} {prospecto.apellido}",
                    prospecto.telefono or "",
                    prospecto.correo_electronico or "",
                    prospecto.estado.replace("_", " ").title()
                ]
            }
            df_info = pd.DataFrame(info_data)
            df_info.to_excel(writer, sheet_name='Información', index=False)

            # Historial de interacciones
            df.to_excel(writer, sheet_name='Historial', index=False)

            # Aplicar formato
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]

                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 60)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)
        return output

    except Exception as e:
        print(f"❌ Error generando Excel de interacciones: {e}")
        import traceback
        traceback.print_exc()
        return None

# ========== ROUTES ==========

@router.get("/prospectos", response_class=HTMLResponse)
async def listar_prospectos(
    request: Request,
    destino: str = Query(None),
    telefono: str = Query(None),
    medio_ingreso_id: str = Query(None),
    agente_asignado_id: str = Query(None),
    estado: str = Query(None),
    busqueda_global: str = Query(None),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)

    if not user:
        return RedirectResponse(url="/", status_code=303)

    # Determinación de fechas
    if not fecha_inicio or not fecha_fin:
        hoy = date.today()
        fecha_inicio_date = date(hoy.year, hoy.month, 1)
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
        fecha_fin_date = date(hoy.year, hoy.month, ultimo_dia)
        fecha_inicio = fecha_inicio_date.strftime("%Y-%m-%d")
        fecha_fin = fecha_fin_date.strftime("%Y-%m-%d")
    else:
        fecha_inicio_date = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        fecha_fin_date = datetime.strptime(fecha_fin, "%Y-%m-%d").date()

    # Query base
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        query = db.query(models.Prospecto).filter(
            models.Prospecto.agente_asignado_id == user.id
        )
    else:
        query = db.query(models.Prospecto)

    # Filtros
    filtros_aplicados = False

    if estado:
        filtros_aplicados = True
        if estado == "todos":
            pass
        else:
            query = query.filter(models.Prospecto.estado == estado)

    if agente_asignado_id:
        filtros_aplicados = True
        if agente_asignado_id == "todos":
            pass
        elif agente_asignado_id == "sin_asignar":
            query = query.filter(models.Prospecto.agente_asignado_id == None)
        else:
            query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))

    # Default filters
    if estado is None and agente_asignado_id is None:
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.Prospecto.estado.in_([
                EstadoProspecto.NUEVO.value,
                EstadoProspecto.EN_SEGUIMIENTO.value,
                EstadoProspecto.COTIZADO.value
            ]))
        else:
            query = query.filter(models.Prospecto.estado == EstadoProspecto.NUEVO.value)
            query = query.filter(models.Prospecto.agente_asignado_id == None)

            estado = EstadoProspecto.NUEVO.value
            agente_asignado_id = "sin_asignar"

    # Filtro fecha
    query = query.filter(
        models.Prospecto.fecha_registro >= fecha_inicio_date,
        models.Prospecto.fecha_registro <= fecha_fin_date
    )

    # Busqueda Global
    if busqueda_global:
        search_term = f"%{busqueda_global}%"
        query = query.filter(
            or_(
                models.Prospecto.nombre.ilike(search_term),
                models.Prospecto.apellido.ilike(search_term),
                models.Prospecto.telefono.ilike(search_term),
                models.Prospecto.telefono_secundario.ilike(search_term),
                models.Prospecto.correo_electronico.ilike(search_term),
                models.Prospecto.destino.ilike(search_term),
                models.Prospecto.ciudad_origen.ilike(search_term),
                models.Prospecto.observaciones.ilike(search_term)
            )
        )

    # Filtro telefono
    if telefono:
        telefono_term = f"%{telefono}%"
        query = query.filter(
            or_(
                models.Prospecto.telefono.ilike(telefono_term),
                models.Prospecto.telefono_secundario.ilike(telefono_term)
            )
        )

    # Otros filtros
    if destino:
        query = query.filter(models.Prospecto.destino.ilike(f"%{destino}%"))

    if medio_ingreso_id and medio_ingreso_id != "todos":
        query = query.filter(models.Prospecto.medio_ingreso_id == int(medio_ingreso_id))

    # Ordenar y paginar
    query = query.order_by(models.Prospecto.fecha_registro.desc())

    total_registros = query.count()
    total_pages = (total_registros + limit - 1) // limit

    if page > total_pages and total_pages > 0:
        page = total_pages

    offset = (page - 1) * limit
    prospectos = query.offset(offset).limit(limit).all()

    agentes = db.query(models.Usuario).filter(
        models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value,
        models.Usuario.activo == 1
    ).all()

    medios_ingreso = db.query(models.MedioIngreso).all()

    return templates.TemplateResponse("prospectos.html", {
        "request": request,
        "prospectos": prospectos,
        "current_user": user,
        "agentes": agentes,
        "medios_ingreso": medios_ingreso,
        "estados_prospecto": [estado.value for estado in EstadoProspecto],
        "filtros_activos": {
            "destino": destino,
            "telefono": telefono,
            "medio_ingreso_id": medio_ingreso_id,
            "agente_asignado_id": agente_asignado_id,
            "estado": estado,
            "busqueda_global": busqueda_global
        },
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_registros": total_registros,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin
    })

@router.post("/prospectos")
async def crear_prospecto(
    request: Request,
    telefono: str = Form(...),
    indicativo_telefono: str = Form("57"),
    medio_ingreso_id: int = Form(...),
    nombre: str = Form(None),
    apellido: str = Form(None),
    correo_electronico: str = Form(None),
    ciudad_origen: str = Form(None),
    destino: str = Form(None),
    fecha_ida: str = Form(None),
    fecha_vuelta: str = Form(None),
    pasajeros_adultos: int = Form(1),
    pasajeros_ninos: int = Form(0),
    pasajeros_infantes: int = Form(0),
    observaciones: str = Form(None),
    empresa_segundo_titular: str = Form(None),
    telefono_secundario: str = Form(None),
    indicativo_telefono_secundario: str = Form("57"),
    forzar_nuevo: bool = Form(False),
    agente_asignado_id: int = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        if not indicativo_telefono.isdigit() or len(indicativo_telefono) > 4:
            return RedirectResponse(url="/prospectos?error=Indicativo principal inválido. Solo números, máximo 4 dígitos", status_code=303)

        if indicativo_telefono_secundario and (not indicativo_telefono_secundario.isdigit() or len(indicativo_telefono_secundario) > 4):
            return RedirectResponse(url="/prospectos?error=Indicativo secundario inválido. Solo números, máximo 4 dígitos", status_code=303)

        # Buscar clientes existentes
        clientes_existentes_principal = db.query(models.Prospecto).filter(
            or_(
                models.Prospecto.telefono == telefono,
                models.Prospecto.telefono_secundario == telefono
            )
        ).all()

        clientes_existentes_secundario = []
        if telefono_secundario:
            clientes_existentes_secundario = db.query(models.Prospecto).filter(
                or_(
                    models.Prospecto.telefono == telefono_secundario,
                    models.Prospecto.telefono_secundario == telefono_secundario
                )
            ).all()

        clientes_existentes_set = set()
        todos_clientes_existentes = []

        for cliente in clientes_existentes_principal:
            if cliente.id not in clientes_existentes_set:
                clientes_existentes_set.add(cliente.id)
                todos_clientes_existentes.append(cliente)

        for cliente in clientes_existentes_secundario:
            if cliente.id not in clientes_existentes_set:
                clientes_existentes_set.add(cliente.id)
                todos_clientes_existentes.append(cliente)

        todos_clientes_existentes.sort(key=lambda x: x.fecha_registro, reverse=True)
        cliente_existente_principal = todos_clientes_existentes[0] if todos_clientes_existentes else None

        if todos_clientes_existentes and not forzar_nuevo:
            # Lógica de confirmación de cliente existente
            todos_ids = [c.id for c in todos_clientes_existentes]
            interacciones_previas = db.query(models.Interaccion).filter(models.Interaccion.prospecto_id.in_(todos_ids)).count()
            documentos_previos = db.query(models.Documento).filter(models.Documento.prospecto_id.in_(todos_ids)).count()
            ultimas_interacciones = db.query(models.Interaccion).filter(models.Interaccion.prospecto_id.in_(todos_ids)).order_by(models.Interaccion.fecha_creacion.desc()).limit(5).all()

            nuevos_datos = {
                "telefono": telefono,
                "indicativo_telefono": indicativo_telefono,
                "nombre": nombre,
                "apellido": apellido,
                "correo_electronico": correo_electronico,
                "ciudad_origen": ciudad_origen,
                "destino": destino,
                "fecha_ida": fecha_ida,
                "fecha_vuelta": fecha_vuelta,
                "pasajeros_adultos": pasajeros_adultos,
                "pasajeros_ninos": pasajeros_ninos,
                "pasajeros_infantes": pasajeros_infantes,
                "medio_ingreso_id": medio_ingreso_id,
                "telefono_secundario": telefono_secundario,
                "indicativo_telefono_secundario": indicativo_telefono_secundario,
                "observaciones": observaciones
            }

            agentes = db.query(models.Usuario).filter(models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value).all()

            return templates.TemplateResponse("confirmar_cliente_existente.html", {
                "request": request,
                "cliente_existente_principal": cliente_existente_principal,
                "registros_previos": todos_clientes_existentes,
                "interacciones_previas": interacciones_previas,
                "documentos_previos": documentos_previos,
                "ultimas_interacciones": ultimas_interacciones,
                "nuevos_datos": nuevos_datos,
                "agentes": agentes
            })

        # Determinar agente
        agente_final_id = None
        if agente_asignado_id and agente_asignado_id != 0:
            agente = db.query(models.Usuario).filter(models.Usuario.id == agente_asignado_id, models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value).first()
            if agente:
                agente_final_id = agente_asignado_id
        elif user.tipo_usuario == TipoUsuario.AGENTE.value:
            agente_final_id = user.id

        # Datos finales
        nombre_final = nombre
        apellido_final = apellido
        email_final = correo_electronico

        if not nombre_final and todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.nombre:
                    nombre_final = cliente.nombre
                    break

        if not apellido_final and todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.apellido:
                    apellido_final = cliente.apellido
                    break

        if not email_final and todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.correo_electronico:
                    email_final = cliente.correo_electronico
                    break

        # Copiar campos adicionales
        fecha_nacimiento_final = None
        numero_identificacion_final = None
        direccion_final = None

        if todos_clientes_existentes:
            for cliente in todos_clientes_existentes:
                if cliente.fecha_nacimiento and not fecha_nacimiento_final:
                    fecha_nacimiento_final = cliente.fecha_nacimiento
                if cliente.numero_identificacion and not numero_identificacion_final:
                    numero_identificacion_final = cliente.numero_identificacion
                if cliente.direccion and not direccion_final:
                    direccion_final = cliente.direccion

        fecha_ida_date = normalizar_fecha_input(fecha_ida)
        fecha_vuelta_date = normalizar_fecha_input(fecha_vuelta)

        telefono_normalizado = normalizar_numero(telefono)
        telefono_secundario_normalizado = normalizar_numero(telefono_secundario) if telefono_secundario else None
        nombre_normalizado = normalizar_texto_mayusculas(nombre_final)
        apellido_normalizado = normalizar_texto_mayusculas(apellido_final)
        ciudad_origen_normalizada = normalizar_texto_mayusculas(ciudad_origen)
        destino_normalizado = normalizar_texto_mayusculas(destino)
        email_normalizado = normalizar_email(email_final)
        empresa_segundo_titular_normalizado = normalizar_texto_mayusculas(empresa_segundo_titular)

        cliente_recurrente = len(todos_clientes_existentes) > 0

        prospecto = models.Prospecto(
            nombre=nombre_normalizado,
            apellido=apellido_normalizado,
            correo_electronico=email_normalizado,
            telefono=telefono_normalizado,
            indicativo_telefono=indicativo_telefono,
            telefono_secundario=telefono_secundario_normalizado,
            indicativo_telefono_secundario=indicativo_telefono_secundario,
            ciudad_origen=ciudad_origen_normalizada,
            destino=destino_normalizado,
            fecha_ida=fecha_ida_date,
            fecha_vuelta=fecha_vuelta_date,
            pasajeros_adultos=pasajeros_adultos,
            pasajeros_ninos=pasajeros_ninos,
            pasajeros_infantes=pasajeros_infantes,
            medio_ingreso_id=medio_ingreso_id,
            observaciones=observaciones,
            agente_asignado_id=agente_final_id,
            cliente_recurrente=cliente_recurrente,
            prospecto_original_id=todos_clientes_existentes[0].id if todos_clientes_existentes else None,
            fecha_nacimiento=fecha_nacimiento_final,
            numero_identificacion=numero_identificacion_final,
            direccion=direccion_final,
            empresa_segundo_titular=empresa_segundo_titular_normalizado
        )

        prospecto.verificar_datos_completos()

        db.add(prospecto)
        db.flush()

        if todos_clientes_existentes and todos_clientes_existentes[0].id_cliente:
            prospecto.id_cliente = todos_clientes_existentes[0].id_cliente
        else:
            prospecto.generar_id_cliente()

        prospecto.generar_id_solicitud()

        db.commit()

        if cliente_recurrente:
            interaccion = models.Interaccion(
                prospecto_id=prospecto.id,
                usuario_id=user.id,
                tipo_interaccion="sistema",
                descripcion=f"Cliente recurrente registrado. Teléfono: {telefono}. Registros previos: {len(todos_clientes_existentes)}",
                estado_anterior=cliente_existente_principal.estado if cliente_existente_principal else None,
                estado_nuevo=EstadoProspecto.NUEVO.value
            )
            db.add(interaccion)
            db.commit()

        mensaje = "Prospecto creado correctamente" + (" (Cliente recurrente)" if cliente_recurrente else "")
        return RedirectResponse(url=f"/prospectos?success={mensaje}", status_code=303)

    except Exception as e:
        db.rollback()
        print(f"❌ Error creating prospect: {e}")
        return RedirectResponse(url="/prospectos?error=Error al crear prospecto", status_code=303)

@router.get("/prospectos/nuevo", response_class=HTMLResponse)
async def nuevo_prospecto_form(request: Request, db: Session = Depends(database.get_db)):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    agentes = db.query(models.Usuario).filter(models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value, models.Usuario.activo == 1).all()
    medios_ingreso = db.query(models.MedioIngreso).all()

    return templates.TemplateResponse("prospecto_form.html", { # Assuming template name based on context
        "request": request,
        "agentes": agentes,
        "medios_ingreso": medios_ingreso,
        "user": user
    })

@router.get("/prospectos/{prospecto_id}/editar")
async def mostrar_editar_prospecto(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
    if not prospecto:
        return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)

    if (user.tipo_usuario == TipoUsuario.AGENTE.value and
        prospecto.agente_asignado_id != user.id):
        return RedirectResponse(url="/prospectos?error=No tiene permisos para editar este prospecto", status_code=303)

    medios_ingreso = db.query(models.MedioIngreso).all()

    return templates.TemplateResponse("editar_prospecto.html", {
        "request": request,
        "prospecto": prospecto,
        "medios_ingreso": medios_ingreso,
        "user": user
    })

@router.post("/prospectos/{prospecto_id}/editar")
async def editar_prospecto(
    request: Request,
    prospecto_id: int,
    telefono: str = Form(...),
    indicativo_telefono: str = Form("57"),
    medio_ingreso_id: int = Form(...),
    nombre: str = Form(None),
    apellido: str = Form(None),
    correo_electronico: str = Form(None),
    ciudad_origen: str = Form(None),
    destino: str = Form(None),
    fecha_ida: str = Form(None),
    fecha_vuelta: str = Form(None),
    pasajeros_adultos: int = Form(1),
    pasajeros_ninos: int = Form(0),
    pasajeros_infantes: int = Form(0),
    observaciones: str = Form(None),
    empresa_segundo_titular: str = Form(None),
    telefono_secundario: str = Form(None),
    indicativo_telefono_secundario: str = Form("57"),
    estado: str = Form(None),
    fecha_nacimiento: str = Form(None),
    numero_identificacion: str = Form(None),
    direccion: str = Form(None),
    origen_solicitud: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        if not indicativo_telefono.isdigit() or len(indicativo_telefono) > 4:
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Indicativo principal inválido" if origen_solicitud == "seguimiento" else "/prospectos?error=Indicativo principal inválido. Solo números, máximo 4 dígitos"
            return RedirectResponse(url=redirect_url, status_code=303)

        if indicativo_telefono_secundario and (not indicativo_telefono_secundario.isdigit() or len(indicativo_telefono_secundario) > 4):
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Indicativo secundario inválido" if origen_solicitud == "seguimiento" else "/prospectos?error=Indicativo secundario inválido. Solo números, máximo 4 dígitos"
            return RedirectResponse(url=redirect_url, status_code=303)

        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Prospecto no encontrado" if origen_solicitud == "seguimiento" else "/prospectos?error=Prospecto no encontrado"
            return RedirectResponse(url=redirect_url, status_code=303)

        if (user.tipo_usuario == TipoUsuario.AGENTE.value and
            prospecto.agente_asignado_id != user.id):
            redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=No tiene permisos" if origen_solicitud == "seguimiento" else "/prospectos?error=No tiene permisos para editar este prospecto"
            return RedirectResponse(url=redirect_url, status_code=303)

        fecha_ida_date = normalizar_fecha_input(fecha_ida)
        fecha_vuelta_date = normalizar_fecha_input(fecha_vuelta)
        fecha_nacimiento_date = normalizar_fecha_input(fecha_nacimiento) if fecha_nacimiento else None

        telefono_normalizado = normalizar_numero(telefono)
        telefono_secundario_normalizado = normalizar_numero(telefono_secundario) if telefono_secundario else None
        nombre_normalizado = normalizar_texto_mayusculas(nombre)
        apellido_normalizado = normalizar_texto_mayusculas(apellido)
        ciudad_origen_normalizada = normalizar_texto_mayusculas(ciudad_origen)
        destino_normalizado = normalizar_texto_mayusculas(destino)
        email_normalizado = normalizar_email(correo_electronico)
        numero_identificacion_normalizado = normalizar_numero(numero_identificacion) if numero_identificacion else None
        direccion_normalizada = normalizar_texto_mayusculas(direccion)
        empresa_segundo_titular_normalizado = normalizar_texto_mayusculas(empresa_segundo_titular)

        if estado == EstadoProspecto.VENTA_CANCELADA.value:
            if prospecto.estado != EstadoProspecto.GANADO.value and prospecto.estado_anterior != EstadoProspecto.GANADO.value:
                redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Solo se puede cancelar una venta que haya estado en estado GANADO" if origen_solicitud == "seguimiento" else "/prospectos?error=Solo se puede cancelar una venta que haya estado en estado GANADO"
                return RedirectResponse(url=redirect_url, status_code=303)

        estado_cambio_a_ganado = False
        if estado and estado != prospecto.estado:
            prospecto.estado_anterior = prospecto.estado
            prospecto.estado = estado
            if estado == EstadoProspecto.GANADO.value:
                estado_cambio_a_ganado = True

        fecha_ida_original = prospecto.fecha_ida
        fecha_ida_cambio = (fecha_ida_date != fecha_ida_original)

        prospecto.nombre = nombre_normalizado
        prospecto.apellido = apellido_normalizado
        prospecto.correo_electronico = email_normalizado
        prospecto.telefono = telefono_normalizado
        prospecto.indicativo_telefono = indicativo_telefono
        prospecto.telefono_secundario = telefono_secundario_normalizado
        prospecto.indicativo_telefono_secundario = indicativo_telefono_secundario
        prospecto.ciudad_origen = ciudad_origen_normalizada
        prospecto.destino = destino_normalizado
        prospecto.fecha_ida = fecha_ida_date
        prospecto.fecha_vuelta = fecha_vuelta_date
        prospecto.pasajeros_adultos = pasajeros_adultos
        prospecto.pasajeros_ninos = pasajeros_ninos
        prospecto.pasajeros_infantes = pasajeros_infantes
        prospecto.medio_ingreso_id = medio_ingreso_id
        prospecto.observaciones = observaciones
        prospecto.empresa_segundo_titular = empresa_segundo_titular_normalizado

        if prospecto.estado == EstadoProspecto.GANADO.value:
            prospecto.fecha_nacimiento = fecha_nacimiento_date
            prospecto.numero_identificacion = numero_identificacion_normalizado
            prospecto.direccion = direccion_normalizada

        db.commit()

        if estado_cambio_a_ganado or (prospecto.estado == EstadoProspecto.GANADO.value and fecha_ida_cambio and fecha_ida_date):
            crear_notificaciones_viaje(prospecto, db)

        if origen_solicitud == "seguimiento":
            return RedirectResponse(url=f"/prospectos/{prospecto_id}/seguimiento?success=Datos actualizados correctamente", status_code=303)
        else:
            return RedirectResponse(url="/prospectos?success=Prospecto actualizado correctamente", status_code=303)

    except Exception as e:
        db.rollback()
        print(f"❌ Error updating prospect: {e}")
        redirect_url = f"/prospectos/{prospecto_id}/seguimiento?error=Error al actualizar" if origen_solicitud == "seguimiento" else "/prospectos?error=Error al actualizar prospecto"
        return RedirectResponse(url=redirect_url, status_code=303)

@router.post("/prospectos/{prospecto_id}/eliminar")
async def eliminar_prospecto(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)

        if (user.tipo_usuario == TipoUsuario.AGENTE.value and
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos?error=No tiene permisos para eliminar este prospecto", status_code=303)

        prospecto.fecha_eliminacion = datetime.now()

        if prospecto.estado not in [EstadoProspecto.GANADO.value, EstadoProspecto.CERRADO_PERDIDO.value]:
            prospecto.estado_anterior = prospecto.estado
            prospecto.estado = "eliminado"

        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="sistema",
            descripcion=f"🗑️ Prospecto marcado como eliminado por {user.username}",
            estado_anterior=prospecto.estado_anterior or prospecto.estado,
            estado_nuevo="eliminado"
        )
        db.add(interaccion)

        db.commit()

        return RedirectResponse(url="/prospectos?success=Prospecto eliminado correctamente", status_code=303)

    except Exception as e:
        db.rollback()
        print(f"❌ Error eliminando prospecto: {e}")
        return RedirectResponse(url="/prospectos?error=Error al eliminar prospecto", status_code=303)

@router.post("/prospectos/{prospecto_id}/asignar")
async def asignar_agente(
    request: Request,
    prospecto_id: int,
    agente_id: int = Form(None),
    destino: str = Form(None),
    telefono: str = Form(None),
    medio_ingreso_id: str = Form(None),
    estado: str = Form(None),
    busqueda_global: str = Form(None),
    agente_filtro_id: str = Form(None),
    fecha_inicio: str = Form(None),
    fecha_fin: str = Form(None),
    periodo: str = Form(None),
    tipo_filtro: str = Form(None),
    valor_filtro: str = Form(None),
    pagina: str = Form("1"),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user or user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        raise HTTPException(status_code=403, detail="No tiene permisos para esta acción")

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")

        if not agente_id or agente_id == 0:
            prospecto.agente_asignado_id = None
            mensaje = "Prospecto desasignado correctamente"
        else:
            agente = db.query(models.Usuario).filter(
                models.Usuario.id == agente_id,
                models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
            ).first()
            if not agente:
                raise HTTPException(status_code=404, detail="Agente no encontrado")

            prospecto.agente_asignado_id = agente_id
            mensaje = f"Agente {agente.username} asignado correctamente"

            notificacion = models.Notificacion(
                usuario_id=agente.id,
                prospecto_id=prospecto.id,
                tipo="asignacion",
                mensaje=f"Te han asignado un nuevo prospecto: {prospecto.nombre} {prospecto.apellido or ''}",
                email_enviado=False
            )
            db.add(notificacion)

            if agente.email:
                asunto = "Nuevo Prospecto Asignado 🚀"
                cuerpo = f"Hola {agente.username},\n\nSe te ha asignado el prospecto {prospecto.nombre} {prospecto.apellido}.\n\nIngresa al sistema para gestionarlo."
                enviado = enviar_notificacion_email(agente.email, asunto, cuerpo)
                notificacion.email_enviado = enviado

        db.commit()

        redirect_url = "/prospectos"
        if tipo_filtro and valor_filtro:
            redirect_url = "/prospectos/filtro"

        params = []
        if destino: params.append(f"destino={destino}")
        if telefono: params.append(f"telefono={telefono}")
        if medio_ingreso_id and medio_ingreso_id != 'todos': params.append(f"medio_ingreso_id={medio_ingreso_id}")
        if estado and estado != 'todos': params.append(f"estado={estado}")
        if busqueda_global: params.append(f"busqueda_global={busqueda_global}")
        if agente_filtro_id and agente_filtro_id != 'todos': params.append(f"agente_asignado_id={agente_filtro_id}")
        if tipo_filtro: params.append(f"tipo_filtro={tipo_filtro}")
        if valor_filtro: params.append(f"valor_filtro={valor_filtro}")
        if fecha_inicio: params.append(f"fecha_inicio={fecha_inicio}")
        if fecha_fin: params.append(f"fecha_fin={fecha_fin}")
        if periodo: params.append(f"periodo={periodo}")
        if pagina and pagina != "1": params.append(f"pagina={pagina}")

        if params:
            redirect_url += "?" + "&".join(params) + f"&success={mensaje}"
        else:
            redirect_url += f"?success={mensaje}"

        return RedirectResponse(url=redirect_url, status_code=303)

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"❌ Error asignando agente: {e}")
        return RedirectResponse(url="/prospectos?error=Error al asignar agente", status_code=303)

@router.get("/prospectos/{prospecto_id}/seguimiento")
async def ver_seguimiento(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
    if not prospecto:
        return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)

    if (user.tipo_usuario == TipoUsuario.AGENTE.value and
        prospecto.agente_asignado_id != user.id):
        return RedirectResponse(url="/prospectos?error=No tiene permisos para ver este prospecto", status_code=303)

    return templates.TemplateResponse("seguimiento_prospecto.html", {
        "request": request,
        "prospecto": prospecto,
        "current_user": user,
        "estados_prospecto": [estado.value for estado in EstadoProspecto]
    })

@router.post("/prospectos/{prospecto_id}/interaccion")
async def registrar_interaccion(
    request: Request,
    prospecto_id: int,
    descripcion: str = Form(...),
    tipo_interaccion: str = Form("general"),
    cambio_estado: str = Form(None),
    fecha_proximo_contacto: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)

        if (user.tipo_usuario == TipoUsuario.AGENTE.value and
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos?error=No tiene permisos para este prospecto", status_code=303)

        if cambio_estado and prospecto.estado:
            estados_orden = [
                EstadoProspecto.NUEVO.value,
                EstadoProspecto.EN_SEGUIMIENTO.value,
                EstadoProspecto.COTIZADO.value,
                EstadoProspecto.GANADO.value,
                EstadoProspecto.CERRADO_PERDIDO.value
            ]

            estado_actual_idx = estados_orden.index(prospecto.estado) if prospecto.estado in estados_orden else -1
            estado_nuevo_idx = estados_orden.index(cambio_estado) if cambio_estado in estados_orden else -1

            if (estado_nuevo_idx < estado_actual_idx and
                estado_actual_idx >= 2 and
                user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]):
                return RedirectResponse(
                    url=f"/prospectos/{prospecto_id}/seguimiento?error=No puede regresar a un estado anterior",
                    status_code=303
                )

        if cambio_estado == EstadoProspecto.CERRADO_PERDIDO.value and not descripcion.strip():
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=Debe agregar un comentario al cerrar el prospecto",
                status_code=303
            )

        if cambio_estado and cambio_estado != prospecto.estado:
            historial = models.HistorialEstado(
                prospecto_id=prospecto_id,
                estado_anterior=prospecto.estado,
                estado_nuevo=cambio_estado,
                usuario_id=user.id,
                comentario=descripcion
            )
            db.add(historial)

        estado_anterior = prospecto.estado
        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion=tipo_interaccion,
            descripcion=descripcion,
            estado_anterior=estado_anterior,
            estado_nuevo=cambio_estado
        )

        db.add(interaccion)

        if fecha_proximo_contacto:
            try:
                fecha_prog = datetime.strptime(fecha_proximo_contacto, "%Y-%m-%dT%H:%M")

                notificacion = models.Notificacion(
                    usuario_id=user.id,
                    prospecto_id=prospecto_id,
                    tipo="seguimiento",
                    mensaje=f"Recordatorio: {descripcion[:50]}...",
                    fecha_programada=fecha_prog,
                    email_enviado=False
                )
                db.add(notificacion)

                if user.email:
                    enviar_notificacion_email(
                        user.email,
                        "Recordatorio Programado 📅",
                        f"Has programado un seguimiento para el prospecto {prospecto.nombre} el {fecha_prog}."
                    )
            except ValueError:
                print(f"❌ Error formato fecha recordatorio: {fecha_proximo_contacto}")

        if (cambio_estado == EstadoProspecto.COTIZADO.value and
            estado_anterior != EstadoProspecto.COTIZADO.value and
            prospecto.agente_asignado_id):

            estadistica = models.EstadisticaCotizacion(
                agente_id=prospecto.agente_asignado_id,
                prospecto_id=prospecto_id,
                fecha_cotizacion=datetime.now().date()
            )
            db.add(estadistica)
            db.flush()
            estadistica.generar_id_cotizacion()
            prospecto.id_cotizacion = estadistica.id_cotizacion

        if cambio_estado:
            prospecto.estado = cambio_estado
            if cambio_estado == EstadoProspecto.GANADO.value and not prospecto.fecha_compra:
                prospecto.fecha_compra = datetime.now().date()

        db.commit()

        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?success=Interacción registrada",
            status_code=303
        )

    except Exception as e:
        db.rollback()
        print(f"❌ Error registrando interacción: {e}")
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?error=Error al registrar interacción",
            status_code=303
        )

@router.get("/documentos/{documento_id}/descargar")
async def descargar_documento(
    documento_id: int,
    request: Request,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    documento = db.query(models.Documento).filter(models.Documento.id == documento_id).first()
    if not documento:
        raise HTTPException(status_code=404, detail="Documento no encontrado")

    # Verificar permisos
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == documento.prospecto_id).first()
        if not prospecto or prospecto.agente_asignado_id != user.id:
            raise HTTPException(status_code=403, detail="No tiene permisos para descargar este documento")

    # Construir ruta completa
    ruta_completa = os.path.join(UPLOAD_DIR, documento.ruta_archivo)

    # Si la ruta guardada ya incluye "uploads/", ajustarla
    if documento.ruta_archivo.startswith("uploads/"):
        ruta_completa = documento.ruta_archivo

    if not os.path.exists(ruta_completa):
        # Intentar buscar solo con el nombre de archivo en uploads raíz (retrocompatibilidad)
        ruta_completa = os.path.join(UPLOAD_DIR, documento.nombre_archivo)
        if not os.path.exists(ruta_completa):
            raise HTTPException(status_code=404, detail="Archivo físico no encontrado")

    return StreamingResponse(
        open(ruta_completa, "rb"),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={documento.nombre_archivo}"}
    )

@router.post("/documentos/{documento_id}/eliminar")
async def eliminar_documento(
    documento_id: int,
    request: Request,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    documento = db.query(models.Documento).filter(models.Documento.id == documento_id).first()
    if not documento:
        # No se puede redirigir fácilmente desde aquí si es llamado por JS, pero asumimos form submit
        return RedirectResponse(url="/prospectos?error=Documento no encontrado", status_code=303)

    prospecto_id = documento.prospecto_id

    # Verificar permisos
    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto or prospecto.agente_asignado_id != user.id:
             return RedirectResponse(url=f"/prospectos/{prospecto_id}/seguimiento?error=No tiene permisos", status_code=303)

    try:
        # Eliminar archivo físico
        ruta_completa = os.path.join(UPLOAD_DIR, documento.ruta_archivo)
        if os.path.exists(ruta_completa):
            os.remove(ruta_completa)

        # Eliminar registro
        db.delete(documento)

        # Registrar interacción
        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="documento",
            descripcion=f"Documento eliminado: {documento.nombre_archivo}",
            estado_anterior=None,
            estado_nuevo=None
        )
        db.add(interaccion)

        db.commit()
        return RedirectResponse(url=f"/prospectos/{prospecto_id}/seguimiento?success=Documento eliminado", status_code=303)

    except Exception as e:
        db.rollback()
        print(f"Error eliminando documento: {e}")
        return RedirectResponse(url=f"/prospectos/{prospecto_id}/seguimiento?error=Error eliminando documento", status_code=303)

@router.post("/prospectos/{prospecto_id}/documento")
async def subir_documento(
    request: Request,
    prospecto_id: int,
    archivo: UploadFile = File(...),
    tipo_documento: str = Form("cotizacion"),
    descripcion: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos?error=Prospecto no encontrado", status_code=303)

        if (user.tipo_usuario == TipoUsuario.AGENTE.value and
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos?error=No tiene permisos para este prospecto", status_code=303)

        allowed_extensions = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
        file_ext = os.path.splitext(archivo.filename)[1].lower()

        if file_ext not in allowed_extensions:
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=Solo se permiten archivos PDF, Office e imágenes",
                status_code=303
            )

        ruta_fecha = obtener_ruta_upload_por_fecha()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"{timestamp}_{archivo.filename}"
        ruta_archivo_completa = os.path.join(ruta_fecha, nombre_archivo)

        with open(ruta_archivo_completa, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)

        ruta_relativa = os.path.relpath(ruta_archivo_completa, UPLOAD_DIR)

        documento = models.Documento(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            nombre_archivo=archivo.filename,
            tipo_documento=tipo_documento,
            ruta_archivo=ruta_relativa,
            descripcion=descripcion
        )

        db.add(documento)
        db.flush()
        documento.generar_id_documento()

        if tipo_documento == "cotizacion":
            estado_anterior = prospecto.estado
            prospecto.estado = EstadoProspecto.COTIZADO.value

            estadistica = models.EstadisticaCotizacion(
                agente_id=prospecto.agente_asignado_id or user.id,
                prospecto_id=prospecto_id,
                fecha_cotizacion=datetime.now().date()
            )
            db.add(estadistica)
            db.flush()
            estadistica.generar_id_cotizacion()
            prospecto.id_cotizacion = estadistica.id_cotizacion

            interaccion = models.Interaccion(
                prospecto_id=prospecto_id,
                usuario_id=user.id,
                tipo_interaccion="documento",
                descripcion=f"Se subió cotización: {archivo.filename}",
                estado_anterior=estado_anterior,
                estado_nuevo=EstadoProspecto.COTIZADO.value
            )
            db.add(interaccion)

        interaccion_doc = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="documento",
            descripcion=f"Documento subido: {archivo.filename} ({tipo_documento})",
            estado_anterior=prospecto.estado,
            estado_nuevo=prospecto.estado
        )
        db.add(interaccion_doc)

        db.commit()

        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?success=Documento subido correctamente",
            status_code=303
        )

    except Exception as e:
        db.rollback()
        print(f"❌ Error subiendo documento: {e}")
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?error=Error al subir documento",
            status_code=303
        )

@router.post("/prospectos/{prospecto_id}/actualizar-viaje")
async def actualizar_viaje(
    request: Request,
    prospecto_id: int,
    nombre: str = Form(None),
    apellido: str = Form(None),
    correo_electronico: str = Form(None),
    telefono: str = Form(...),
    indicativo_telefono: str = Form("57"),
    indicativo_telefono_secundario: str = Form("57"),
    ciudad_origen: str = Form(None),
    destino: str = Form(None),
    fecha_ida: str = Form(None),
    fecha_vuelta: str = Form(None),
    pasajeros_adultos: int = Form(1),
    pasajeros_ninos: int = Form(0),
    pasajeros_infantes: int = Form(0),
    telefono_secundario: str = Form(None),
    fecha_nacimiento: str = Form(None),
    numero_identificacion: str = Form(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=Prospecto no encontrado",
                status_code=303
            )

        if (user.tipo_usuario == models.TipoUsuario.AGENTE.value and
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(
                url=f"/prospectos/{prospecto_id}/seguimiento?error=No tiene permisos para editar este prospecto",
                status_code=303
            )

        fecha_ida_date = parsear_fecha(fecha_ida)
        fecha_vuelta_date = parsear_fecha(fecha_vuelta)

        prospecto.nombre = nombre
        prospecto.apellido = apellido
        prospecto.correo_electronico = correo_electronico
        prospecto.ciudad_origen = ciudad_origen
        prospecto.destino = destino
        prospecto.telefono = telefono
        prospecto.indicativo_telefono = indicativo_telefono
        prospecto.telefono_secundario = telefono_secundario
        prospecto.indicativo_telefono_secundario = indicativo_telefono_secundario
        prospecto.fecha_ida = fecha_ida_date
        prospecto.fecha_vuelta = fecha_vuelta_date
        prospecto.pasajeros_adultos = pasajeros_adultos
        prospecto.pasajeros_ninos = pasajeros_ninos
        prospecto.pasajeros_infantes = pasajeros_infantes

        fecha_nacimiento_date = parsear_fecha(fecha_nacimiento) if fecha_nacimiento else None
        if prospecto.estado == EstadoProspecto.GANADO.value or prospecto.estado == EstadoProspecto.VENTA_CANCELADA.value:
            prospecto.fecha_nacimiento = fecha_nacimiento_date
            prospecto.numero_identificacion = numero_identificacion

        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="sistema",
            descripcion="Información de viaje actualizada",
            estado_anterior=prospecto.estado,
            estado_nuevo=prospecto.estado
        )

        db.add(interaccion)
        db.commit()

        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?success=Información de viaje actualizada correctamente",
            status_code=303
        )

    except Exception as e:
        db.rollback()
        print(f"❌ Error actualizando viaje: {e}")
        return RedirectResponse(
            url=f"/prospectos/{prospecto_id}/seguimiento?error=Error al actualizar información",
            status_code=303
        )

@router.get("/prospectos/cerrados", response_class=HTMLResponse)
async def listar_prospectos_cerrados(
    request: Request,
    fecha_registro_desde: str = Query(None),
    fecha_registro_hasta: str = Query(None),
    fecha_cierre_desde: str = Query(None),
    fecha_cierre_hasta: str = Query(None),
    destino: str = Query(None),
    agente_asignado_id: str = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(10, ge=1, le=100),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    query = db.query(models.Prospecto).filter(
        models.Prospecto.estado.in_([EstadoProspecto.CERRADO_PERDIDO.value, EstadoProspecto.GANADO.value])
    )

    if user.tipo_usuario == TipoUsuario.AGENTE.value:
        query = query.filter(models.Prospecto.agente_asignado_id == user.id)

    if fecha_registro_desde:
        try:
            fecha_desde = datetime.strptime(fecha_registro_desde, "%d/%m/%Y").date()
            query = query.filter(models.Prospecto.fecha_registro >= fecha_desde)
        except ValueError:
            pass

    if fecha_registro_hasta:
        try:
            fecha_hasta = datetime.strptime(fecha_registro_hasta, "%d/%m/%Y").date()
            query = query.filter(models.Prospecto.fecha_registro <= fecha_hasta)
        except ValueError:
            pass

    if fecha_cierre_desde or fecha_cierre_hasta:
        subquery = db.query(models.Interaccion.prospecto_id).filter(
            models.Interaccion.estado_nuevo.in_([EstadoProspecto.CERRADO_PERDIDO.value, EstadoProspecto.GANADO.value])
        )
        if fecha_cierre_desde:
            try:
                fecha_cierre_desde_date = datetime.strptime(fecha_cierre_desde, "%d/%m/%Y").date()
                subquery = subquery.filter(models.Interaccion.fecha_creacion >= fecha_cierre_desde_date)
            except ValueError:
                pass
        if fecha_cierre_hasta:
            try:
                fecha_cierre_hasta_date = datetime.strptime(fecha_cierre_hasta, "%d/%m/%Y").date()
                subquery = subquery.filter(models.Interaccion.fecha_creacion <= fecha_cierre_hasta_date)
            except ValueError:
                pass
        query = query.filter(models.Prospecto.id.in_(subquery))

    if destino:
        search_term = f"%{destino}%"
        query = query.filter(or_(
            models.Prospecto.destino.ilike(search_term),
            models.Prospecto.nombre.ilike(search_term),
            models.Prospecto.apellido.ilike(search_term),
            models.Prospecto.correo_electronico.ilike(search_term),
            models.Prospecto.telefono.ilike(search_term)
        ))

    if agente_asignado_id and agente_asignado_id != "todos" and user.tipo_usuario in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))

    query = query.order_by(models.Prospecto.fecha_registro.desc())

    total_registros = query.count()
    total_pages = (total_registros + limit - 1) // limit

    if page > total_pages and total_pages > 0:
        page = total_pages
    if page < 1:
        page = 1

    offset = (page - 1) * limit
    prospectos_cerrados = query.offset(offset).limit(limit).all()

    agentes = db.query(models.Usuario).filter(models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value).all()

    return templates.TemplateResponse("prospectos_cerrados.html", {
        "request": request,
        "prospectos": prospectos_cerrados,
        "current_user": user,
        "agentes": agentes,
        "filtros_activos": {
            "fecha_registro_desde": fecha_registro_desde,
            "fecha_registro_hasta": fecha_registro_hasta,
            "fecha_cierre_desde": fecha_cierre_desde,
            "fecha_cierre_hasta": fecha_cierre_hasta,
            "destino": destino,
            "agente_asignado_id": agente_asignado_id
        },
        "page": page,
        "limit": limit,
        "total_pages": total_pages,
        "total_registros": total_registros
    })

@router.post("/prospectos/{prospecto_id}/reactivar")
async def reactivar_prospecto(
    request: Request,
    prospecto_id: int,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()
        if not prospecto:
            return RedirectResponse(url="/prospectos/cerrados?error=Prospecto no encontrado", status_code=303)

        if (user.tipo_usuario == TipoUsuario.AGENTE.value and
            prospecto.agente_asignado_id != user.id):
            return RedirectResponse(url="/prospectos/cerrados?error=No tiene permisos para reactivar este prospecto", status_code=303)

        estado_anterior = prospecto.estado
        prospecto.estado = EstadoProspecto.EN_SEGUIMIENTO.value

        interaccion = models.Interaccion(
            prospecto_id=prospecto_id,
            usuario_id=user.id,
            tipo_interaccion="sistema",
            descripcion=f"Prospecto reactivado desde estado: {estado_anterior}",
            estado_anterior=estado_anterior,
            estado_nuevo=EstadoProspecto.EN_SEGUIMIENTO.value
        )

        db.add(interaccion)
        db.commit()

        return RedirectResponse(url="/prospectos/cerrados?success=Prospecto reactivado correctamente", status_code=303)

    except Exception as e:
        db.rollback()
        print(f"❌ Error reactivando prospecto: {e}")
        return RedirectResponse(url="/prospectos/cerrados?error=Error al reactivar prospecto", status_code=303)

@router.get("/prospectos/exportar/excel")
async def exportar_prospectos_excel(
    request: Request,
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(get_current_user)
):
    try:
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.agente_asignado_id == user.id
            ).all()
        else:
            prospectos = db.query(models.Prospecto).all()

        data = []
        for p in prospectos:
            data.append({
                'ID': p.id,
                'Nombre': f"{p.nombre or ''} {p.apellido or ''}",
                'Email': p.correo_electronico or '',
                'Teléfono': p.telefono or '',
                'Destino': p.destino or '',
                'Estado': p.estado,
                'Agente': p.agente_asignado.username if p.agente_asignado else 'Sin asignar',
                'Fecha Registro': p.fecha_registro.strftime('%d/%m/%Y'),
                'Medio Ingreso': p.medio_ingreso.nombre
            })

        df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Prospectos', index=False)

        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=prospectos.xlsx"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exportando: {str(e)}")

@router.get("/prospectos/filtro", response_class=HTMLResponse)
async def prospectos_filtro_dashboard(
    request: Request,
    tipo_filtro: str = Query(...),
    valor_filtro: str = Query(...),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    periodo: str = Query("mes"),
    pagina: int = Query(1),
    agente_asignado_id: str = Query(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)
    if not user:
        return RedirectResponse(url="/", status_code=303)

    registros_por_pagina = 50
    offset = (pagina - 1) * registros_por_pagina

    fecha_inicio_dt, fecha_fin_dt = calcular_rango_fechas(periodo, fecha_inicio, fecha_fin)
    fecha_inicio_date = fecha_inicio_dt.date()
    fecha_fin_date = fecha_fin_dt.date()

    if tipo_filtro == "estado" and valor_filtro == EstadoProspecto.COTIZADO.value:
        query = db.query(models.Prospecto).join(
            models.EstadisticaCotizacion,
            models.EstadisticaCotizacion.prospecto_id == models.Prospecto.id
        ).filter(
            models.EstadisticaCotizacion.fecha_cotizacion >= fecha_inicio_date,
            models.EstadisticaCotizacion.fecha_cotizacion <= fecha_fin_date
        )
        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.EstadisticaCotizacion.agente_id == user.id)

        titulo_filtro = "Prospectos que han sido cotizados en el periodo"

    elif (tipo_filtro == "estado" and valor_filtro in [
            EstadoProspecto.EN_SEGUIMIENTO.value,
            EstadoProspecto.GANADO.value,
            EstadoProspecto.CERRADO_PERDIDO.value
        ]) or (tipo_filtro == "ventas"):

        target_state = valor_filtro
        if tipo_filtro == "ventas":
            target_state = EstadoProspecto.GANADO.value
            titulo_filtro = "Ventas realizadas en el periodo"
        else:
            titulo_filtro = f"Prospectos {valor_filtro.replace('_', ' ').title()} en el periodo"

        query = db.query(models.Prospecto).join(
            models.HistorialEstado,
            models.HistorialEstado.prospecto_id == models.Prospecto.id
        ).filter(
            models.HistorialEstado.estado_nuevo == target_state,
            models.HistorialEstado.fecha_cambio >= fecha_inicio_dt,
            models.HistorialEstado.fecha_cambio <= fecha_fin_dt
        )

        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.HistorialEstado.usuario_id == user.id)

    else:
        query = db.query(models.Prospecto).filter(
            models.Prospecto.fecha_registro >= fecha_inicio_dt,
            models.Prospecto.fecha_registro <= fecha_fin_dt
        )

        if user.tipo_usuario == TipoUsuario.AGENTE.value:
            query = query.filter(models.Prospecto.agente_asignado_id == user.id)

    if agente_asignado_id and agente_asignado_id != "todos":
        try:
            if tipo_filtro == "estado" and valor_filtro == EstadoProspecto.COTIZADO.value:
                query = query.filter(models.EstadisticaCotizacion.agente_id == int(agente_asignado_id))
            elif (tipo_filtro == "estado" and valor_filtro in [EstadoProspecto.EN_SEGUIMIENTO.value, EstadoProspecto.GANADO.value, EstadoProspecto.CERRADO_PERDIDO.value]) or (tipo_filtro == "ventas"):
                query = query.filter(models.HistorialEstado.usuario_id == int(agente_asignado_id))
            else:
                query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))
        except ValueError:
            pass

    if tipo_filtro == "estado" and valor_filtro != EstadoProspecto.COTIZADO.value:
        query = query.filter(models.Prospecto.estado == valor_filtro)
        titulo_filtro = f"Prospectos en estado: {valor_filtro.replace('_', ' ').title()}"

    elif tipo_filtro == "asignacion":
        if valor_filtro == "sin_asignar":
            query = query.filter(models.Prospecto.agente_asignado_id == None)
            titulo_filtro = "Prospectos sin asignar"
        elif valor_filtro == "asignados":
            query = query.filter(models.Prospecto.agente_asignado_id != None)
            titulo_filtro = "Prospectos asignados"

    elif tipo_filtro == "destino":
        query = query.filter(models.Prospecto.destino.ilike(f"%{valor_filtro}%"))
        titulo_filtro = f"Prospectos con destino: {valor_filtro}"

    elif tipo_filtro == "datos":
        if valor_filtro == "con_datos":
            query = query.filter(models.Prospecto.tiene_datos_completos == True)
            titulo_filtro = "Prospectos con datos completos"
        elif valor_filtro == "sin_datos":
            query = query.filter(models.Prospecto.tiene_datos_completos == False)
            titulo_filtro = "Prospectos sin datos (solo teléfono)"

    elif tipo_filtro == "total":
        titulo_filtro = "Todos los prospectos registrados"

    total_prospectos = query.count()
    prospectos = query.offset(offset).limit(registros_por_pagina).all()

    total_paginas = (total_prospectos + registros_por_pagina - 1) // registros_por_pagina

    agentes = db.query(models.Usuario).filter(
        models.Usuario.tipo_usuario == TipoUsuario.AGENTE.value
    ).all()

    medios_ingreso = db.query(models.MedioIngreso).all()

    return templates.TemplateResponse("prospectos_filtro.html", {
        "request": request,
        "prospectos": prospectos,
        "current_user": user,
        "agentes": agentes,
        "medios_ingreso": medios_ingreso,
        "titulo_filtro": titulo_filtro,
        "tipo_filtro": tipo_filtro,
        "valor_filtro": valor_filtro,
        "pagina_actual": pagina,
        "total_paginas": total_paginas,
        "total_prospectos": total_prospectos,
        "registros_por_pagina": registros_por_pagina,
        "fecha_inicio_activa": fecha_inicio,
        "fecha_fin_activa": fecha_fin,
        "periodo_activo": periodo,
        "fecha_inicio_formateada": fecha_inicio_dt.strftime("%d/%m/%Y"),
        "fecha_fin_formateada": fecha_fin_dt.strftime("%d/%m/%Y")
    })

@router.get("/api/destinos/sugerencias")
async def sugerencias_destinos(
    q: str = Query("", min_length=2),
    limit: int = Query(10),
    db: Session = Depends(database.get_db)
):
    if len(q) < 2:
        return JSONResponse(content={"sugerencias": []})

    try:
        destinos = db.query(models.Prospecto.destino).filter(
            models.Prospecto.destino.isnot(None),
            models.Prospecto.destino != '',
            models.Prospecto.destino.ilike(f"%{q}%")
        ).distinct().limit(limit).all()

        sugerencias = [destino[0] for destino in destinos if destino[0]]

        sugerencias.sort(key=lambda x:
            0 if x.lower().startswith(q.lower()) else
            1 if q.lower() in x.lower() else 2
        )

        return JSONResponse(content={"sugerencias": sugerencias[:limit]})

    except Exception as e:
        print(f"Error en sugerencias_destinos: {e}")
        return JSONResponse(content={"sugerencias": []})

@router.post("/api/destinos/normalizar")
async def normalizar_destinos(
    destino_original: str = Form(...),
    destino_normalizado: str = Form(...),
    aplicar_a_todos: bool = Form(False),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(get_current_user)
):
    if not user or user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
        raise HTTPException(status_code=403, detail="No tiene permisos")

    try:
        if aplicar_a_todos:
            prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.destino.ilike(f"%{destino_original}%")
            ).all()

            for prospecto in prospectos:
                prospecto.destino = destino_normalizado

            count = len(prospectos)
            mensaje = f"Se normalizaron {count} prospectos"
        else:
            prospectos = db.query(models.Prospecto).filter(
                models.Prospecto.destino == destino_original
            ).all()

            for prospecto in prospectos:
                prospecto.destino = destino_normalizado

            count = len(prospectos)
            mensaje = f"Se normalizaron {count} prospectos"

        db.commit()

        if count > 0:
            accion = models.Interaccion(
                prospecto_id=None,
                usuario_id=user.id,
                tipo_interaccion="sistema",
                descripcion=f"Normalización de destinos: '{destino_original}' → '{destino_normalizado}' ({count} registros)",
                estado_anterior=None,
                estado_nuevo=None
            )
            db.add(accion)
            db.commit()

        return {"success": True, "message": mensaje, "count": count}

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")

@router.post("/importar-prospectos")
async def importar_prospectos(
    request: Request,
    archivo: UploadFile = File(...),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    try:
        es_valido, mensaje_error = excel_import.validar_archivo_excel(archivo)
        if not es_valido:
            return templates.TemplateResponse("importar_datos.html", {
                "request": request,
                "current_user": user,
                "resultado_prospectos": {
                    "exitosos": 0,
                    "errores": [{"fila": 0, "error": mensaje_error}],
                    "recurrentes": 0
                }
            })

        temp_path = f"uploads/temp_{archivo.filename}"
        with open(temp_path, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)

        resultado = excel_import.importar_prospectos_desde_excel(temp_path, db)

        os.remove(temp_path)

        return templates.TemplateResponse("importar_datos.html", {
            "request": request,
            "current_user": user,
            "resultado_prospectos": resultado
        })

    except Exception as e:
        return templates.TemplateResponse("importar_datos.html", {
            "request": request,
            "current_user": user,
            "resultado_prospectos": {
                "exitosos": 0,
                "errores": [{"fila": 0, "error": f"Error procesando archivo: {str(e)}"}],
                "recurrentes": 0
            }
        })

@router.get("/exportar/prospectos")
async def exportar_prospectos(
    request: Request,
    destino: str = Query(None),
    telefono: str = Query(None),
    medio_ingreso_id: str = Query(None),
    agente_asignado_id: str = Query(None),
    estado: str = Query(None),
    busqueda_global: str = Query(None),
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)

    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        query = db.query(models.Prospecto).filter(
            models.Prospecto.fecha_eliminacion.is_(None)
        )

        if user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
            query = query.filter(models.Prospecto.agente_asignado_id == user.id)

        if destino:
            query = query.filter(models.Prospecto.destino.ilike(f"%{destino}%"))

        if telefono:
            telefono_normalizado = normalizar_numero(telefono)
            query = query.filter(
                or_(
                    func.replace(func.replace(models.Prospecto.telefono, ' ', ''), '-', '').ilike(f"%{telefono_normalizado}%"),
                    func.replace(func.replace(models.Prospecto.telefono_secundario, ' ', ''), '-', '').ilike(f"%{telefono_normalizado}%")
                )
            )

        if medio_ingreso_id and medio_ingreso_id != "todos":
            query = query.filter(models.Prospecto.medio_ingreso_id == int(medio_ingreso_id))

        if agente_asignado_id and agente_asignado_id != "todos":
            if agente_asignado_id == "sin_asignar":
                query = query.filter(models.Prospecto.agente_asignado_id.is_(None))
            else:
                query = query.filter(models.Prospecto.agente_asignado_id == int(agente_asignado_id))

        if estado and estado != "todos":
            query = query.filter(models.Prospecto.estado == estado)

        if busqueda_global:
            busqueda = f"%{busqueda_global}%"
            query = query.filter(
                or_(
                    models.Prospecto.nombre.ilike(busqueda),
                    models.Prospecto.apellido.ilike(busqueda),
                    models.Prospecto.telefono.ilike(busqueda),
                    models.Prospecto.correo_electronico.ilike(busqueda),
                    models.Prospecto.id_cliente.ilike(busqueda)
                )
            )

        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
                fecha_inicio_dt = datetime.combine(fecha_inicio_obj, datetime.min.time())
                query = query.filter(models.Prospecto.fecha_registro >= fecha_inicio_dt)
            except ValueError:
                pass

        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, "%d/%m/%Y").date()
                fecha_fin_dt = datetime.combine(fecha_fin_obj, datetime.max.time())
                query = query.filter(models.Prospecto.fecha_registro <= fecha_fin_dt)
            except ValueError:
                pass

        prospectos = query.order_by(models.Prospecto.fecha_registro.desc()).limit(10000).all()

        excel_file = generar_excel_prospectos(prospectos)

        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"prospectos_{timestamp}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Error exportando prospectos: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar datos")

@router.get("/exportar/interacciones/{prospecto_id}")
async def exportar_interacciones(
    prospecto_id: int,
    request: Request,
    db: Session = Depends(database.get_db)
):
    user = await get_current_user(request, db)

    if not user:
        return RedirectResponse(url="/", status_code=303)

    try:
        prospecto = db.query(models.Prospecto).filter(models.Prospecto.id == prospecto_id).first()

        if not prospecto:
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")

        if user.tipo_usuario not in [TipoUsuario.ADMINISTRADOR.value, TipoUsuario.SUPERVISOR.value]:
            if prospecto.agente_asignado_id != user.id:
                raise HTTPException(status_code=403, detail="No tiene permiso para ver este prospecto")

        interacciones = db.query(models.Interaccion).filter(
            models.Interaccion.prospecto_id == prospecto_id
        ).order_by(models.Interaccion.fecha_creacion.desc()).all()

        excel_file = generar_excel_interacciones(interacciones, prospecto)

        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        id_cliente = prospecto.id_cliente or f"CL-{prospecto.id:04d}"
        filename = f"interacciones_{id_cliente}_{timestamp}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error exportando interacciones: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar interacciones")

@router.get("/exportar/clientes-ganados")
async def exportar_clientes_ganados(
    request: Request,
    fecha_inicio: str = Query(None),
    fecha_fin: str = Query(None),
    db: Session = Depends(database.get_db),
    user: models.Usuario = Depends(require_admin)
):
    try:
        query = db.query(models.Prospecto).filter(
            models.Prospecto.estado == EstadoProspecto.GANADO.value,
            models.Prospecto.fecha_eliminacion.is_(None)
        )

        if fecha_inicio:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, "%d/%m/%Y").date()
                fecha_inicio_dt = datetime.combine(fecha_inicio_obj, datetime.min.time())
                query = query.filter(models.Prospecto.fecha_compra >= fecha_inicio_dt)
            except ValueError:
                pass

        if fecha_fin:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, "%d/%m/%Y").date()
                fecha_fin_dt = datetime.combine(fecha_fin_obj, datetime.max.time())
                query = query.filter(models.Prospecto.fecha_compra <= fecha_fin_dt)
            except ValueError:
                pass

        clientes = query.order_by(models.Prospecto.fecha_compra.desc()).all()

        excel_file = generar_excel_prospectos(clientes)

        if not excel_file:
            raise HTTPException(status_code=500, detail="Error generando archivo Excel")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"clientes_ganados_{timestamp}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        print(f"❌ Error exportando clientes ganados: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error al exportar clientes ganados")
